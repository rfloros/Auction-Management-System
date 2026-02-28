from dataclasses import dataclass, field
from fileinput import filename
from AuctionClasses.Bidders import Bidder
from AuctionClasses.Items import Item
import json
from openpyxl import Workbook, load_workbook



@dataclass
class Auction:
    bidders: dict[int, Bidder] = field(default_factory=dict)
    items: dict[int, Item] = field(default_factory=dict)

    # ---------- Internal helpers ----------
    def _getItem(self, itemNumber: int) -> Item:
        if itemNumber not in self.items:
            raise ValueError(f"Item {itemNumber} does not exist")
        return self.items[itemNumber]

    def _getBidder(self, bidderId: int) -> Bidder:
        if bidderId not in self.bidders:
            raise ValueError(f"Bidder {bidderId} does not exist")
        return self.bidders[bidderId]

    def _ensureItemDoesNotExist(self, itemNumber: int):
        if itemNumber in self.items:
            raise ValueError(f"Item {itemNumber} already exists")

    def _ensureBidderDoesNotExist(self, bidderId: int):
        if bidderId in self.bidders:
            raise ValueError(f"Bidder {bidderId} already exists")

    # ---------- Setup ----------
    def addItem(self, itemNumber: int, name: str, type: str):
        self._ensureItemDoesNotExist(itemNumber)
        self.items[itemNumber] = Item(itemNumber, name, type)

    def checkInBidder(self, bidderId: int, name: str):
        self._ensureBidderDoesNotExist(bidderId)
        self.bidders[bidderId] = Bidder(bidderId, name)

    # ---------- Core transaction ----------
    def recordSale(self, itemNumber: int, bidderId: int, salePrice: float):

        item = self._getItem(itemNumber)
        bidder = self._getBidder(bidderId)

        if item.winnerId is not None:
            raise ValueError("Item already sold")

        if salePrice <= 0:
            raise ValueError("Sale price must be positive")

        item.salePrice = salePrice
        item.winnerId = bidderId

        bidder.itemsWon.append(itemNumber)
        bidder.totalOwed += salePrice

    # ---------- Data reports ----------
    def getBidderReceipt(self, bidderId: int) -> str:
        bidder = self._getBidder(bidderId)

        lines = [f"Receipt for {bidder.name} (Bidder ID: {bidder.bidderId})"]
        lines.append("Items Won:")

        if not bidder.itemsWon:
            lines.append(" - None")

        for itemNumber in bidder.itemsWon:
            item = self._getItem(itemNumber)
            lines.append(f" - {item.name} (Item {item.itemNumber}): ${item.salePrice:.2f}")

        lines.append(f"Total Owed: ${bidder.totalOwed:.2f}")
        return "\n".join(lines)

    def getAuctionSummary(self) -> str:
        lines = ["Auction Summary:"]

        for item in self.items.values():
            if item.winnerId is None:
                lines.append(f" - {item.name} (Item {item.itemNumber}): Not Sold")
            else:
                bidder = self._getBidder(item.winnerId)
                lines.append(
                    f" - {item.name} (Item {item.itemNumber}): Sold to {bidder.name} for ${item.salePrice:.2f}"
                )

        totalRevenue = self.getTotalRevenue()
        lines.append(f"\nTotal Revenue: ${totalRevenue:.2f}")

        return "\n".join(lines)
    
    def getTotalRevenue(self) -> float:
        total = 0.0
        for item in self.items.values():
            if item.salePrice is not None:
                total += item.salePrice
        return total

    # ---------- Printable helpers ----------
    def printBidderReceipt(self, bidderId: int):
        print(self.getBidderReceipt(bidderId))

    def printAuctionSummary(self):
        print(self.getAuctionSummary())

    # ---------- Data save and load ----------
    def saveToFile(self, filename: str):
        data = {
            "bidders": {bidderId: bidder.to_dict() for bidderId, bidder in self.bidders.items()},
            "items": {itemNumber: item.to_dict() for itemNumber, item in self.items.items()}
        }
        with open(filename, "w") as f:
            json.dump(data, f, indent=4)

    def loadFromFile(self, filename: str):
        with open(filename, "r") as file:
            data = json.load(file)

        self.items.clear()
        self.bidders.clear()

        for id, itemData in data["items"].items():
            item = Item(
                itemNumber=itemData["itemNumber"],
                name=itemData["name"],
                type=itemData["type"],
                salePrice=itemData["salePrice"],
                winnerId=itemData["winnerId"]
        )
        self.items[int(id)] = item

        for id, bidderData in data["bidders"].items():
            bidder = Bidder(
                bidderId=bidderData["bidderId"],
                name=bidderData["name"],
                itemsWon=bidderData["itemsWon"],
                totalOwed=bidderData["totalOwed"]
        )
        self.bidders[int(id)] = bidder

        #---------- Excel save and load ----------
    def saveToExcel(self, filename: str):
        wb = Workbook()

        # Create sheets
        itemsSheet = wb.active
        itemsSheet.title = "Items"
        biddersSheet = wb.create_sheet("Bidders")

        # Write headers
        itemsSheet.append(["ItemId", "Name", "Type", "SalePrice", "WinnerId"])
        biddersSheet.append(["BidderId", "Name", "TotalOwed"])

        # Write items
        for item in self.items.values():
            itemsSheet.append([
                item.itemNumber,
                item.name,
                item.type,
                item.salePrice,
                item.winnerId
            ])

        # Write bidders
        for bidder in self.bidders.values():
            biddersSheet.append([
                bidder.bidderId,
                bidder.name,
                bidder.totalOwed
            ])

        wb.save(filename)

    def loadFromExcel(self, filename: str):
        wb = load_workbook(filename)

        itemsSheet = wb["Items"]
        biddersSheet = wb["Bidders"]

        self.items.clear()
        self.bidders.clear()

        # Load items
        for row in itemsSheet.iter_rows(min_row=2, values_only=True):
            itemNumber, name, type, salePrice, winnerId = row
            item = Item(itemNumber, name, type, salePrice, winnerId)
            self.items[itemNumber] = item

        # Load bidders
        for row in biddersSheet.iter_rows(min_row=2, values_only=True):
            bidderId, name, totalOwed = row
            bidder = Bidder(bidderId, name)
            bidder.totalOwed = totalOwed
            self.bidders[bidderId] = bidder

        # Reset bidder state
        for bidder in self.bidders.values():
            bidder.totalOwed = 0
            bidder.itemsWon.clear()

        # Rebuild relationships
        for item in self.items.values():
            if item.winnerId is not None:
                bidder = self.bidders.get(item.winnerId)
                if bidder:
                    bidder.itemsWon.append(item.itemNumber)
                    if item.salePrice is not None:
                        bidder.totalOwed += item.salePrice